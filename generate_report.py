#!/usr/bin/env python3
"""
Generates:
  docs/index.html   — GitHub Pages live dashboard
  data/trades.csv   — running trade log, appended each cycle
  data/trades.xlsx  — formatted Excel workbook, rebuilt each cycle
"""
import csv
import json
import os
from datetime import datetime

DATA_DIR = "data"
DOCS_DIR = "docs"

AGENT_NAMES = [
    "Momentum Agent",
    "RSI Agent",
    "MA Crossover",
    "Dividend Agent",
    "Sentiment Agent",
]
COLORS = {
    "Momentum Agent":  "#f0883e",
    "RSI Agent":       "#58a6ff",
    "MA Crossover":    "#3fb950",
    "Dividend Agent":  "#d2a8ff",
    "Sentiment Agent": "#ffa657",
}


# ─────────────────────────────────────────────────────────────────────────────
# Data helpers
# ─────────────────────────────────────────────────────────────────────────────

def load_portfolio(name: str) -> dict:
    path = os.path.join(DATA_DIR, f"{name.lower().replace(' ', '_')}.json")
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return {}


def load_snapshots() -> list:
    path = os.path.join(DATA_DIR, "snapshots.json")
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return []


def save_snapshot(portfolios: dict):
    path = os.path.join(DATA_DIR, "snapshots.json")
    snapshots = load_snapshots()
    entry = {"time": datetime.utcnow().strftime("%Y-%m-%d %H:%M")}
    for name, pf in portfolios.items():
        if pf:
            pos_val = sum(p["shares"] * p["avg_price"] for p in pf.get("positions", {}).values())
            entry[name] = round(pf["cash"] + pos_val, 2)
    snapshots.append(entry)
    snapshots = snapshots[-1000:]
    with open(path, "w") as f:
        json.dump(snapshots, f)


# ─────────────────────────────────────────────────────────────────────────────
# CSV — updated in place, never recreated from scratch
# ─────────────────────────────────────────────────────────────────────────────

CSV_PATH = os.path.join(DATA_DIR, "trades.csv")
CSV_FIELDS = ["time", "agent", "action", "symbol", "shares", "price", "total", "trade_pnl", "reason"]


def update_csv(portfolios: dict):
    # Load existing rows so we can skip already-written trades
    existing_keys: set = set()
    if os.path.exists(CSV_PATH):
        with open(CSV_PATH, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                existing_keys.add((row["time"], row["agent"], row["action"], row["symbol"]))

    new_rows = []
    for name, pf in portfolios.items():
        for t in pf.get("trades", []):
            ts = t["time"][:16].replace("T", " ")
            key = (ts, name, t["action"], t["symbol"])
            if key not in existing_keys:
                new_rows.append({
                    "time":      ts,
                    "agent":     name,
                    "action":    t["action"],
                    "symbol":    t["symbol"],
                    "shares":    int(t["shares"]),
                    "price":     round(t["price"], 4),
                    "total":     round(t["total"], 2),
                    "trade_pnl": t.get("trade_pnl", ""),
                    "reason":    t.get("reason", ""),
                })

    if not new_rows and os.path.exists(CSV_PATH):
        return  # nothing new to write

    write_header = not os.path.exists(CSV_PATH)
    new_rows.sort(key=lambda r: r["time"])

    with open(CSV_PATH, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        if write_header:
            writer.writeheader()
        writer.writerows(new_rows)

    print(f"CSV updated — {len(new_rows)} new row(s) appended → {CSV_PATH}")


# ─────────────────────────────────────────────────────────────────────────────
# Excel (.xlsx) — fully formatted, rebuilt each cycle
# ─────────────────────────────────────────────────────────────────────────────

XLSX_PATH = os.path.join(DATA_DIR, "trades.xlsx")

AGENT_COLORS_HEX = {
    "Momentum Agent":  "F0883E",
    "RSI Agent":       "58A6FF",
    "MA Crossover":    "3FB950",
    "Dividend Agent":  "D2A8FF",
    "Sentiment Agent": "FFA657",
}


def rebuild_xlsx(portfolios: dict):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed — skipping xlsx")
        return

    wb = Workbook()

    # ── Sheet 1: All Trades ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "All Trades"

    headers = ["Time", "Agent", "Action", "Stock", "Shares", "Price ($)", "Total ($)", "Trade P&L ($)", "Why the agent did this"]
    col_widths = [18, 18, 8, 8, 8, 11, 12, 14, 80]

    header_fill  = PatternFill("solid", fgColor="1C2128")
    header_font  = Font(bold=True, color="E6EDF3", size=10)
    buy_fill     = PatternFill("solid", fgColor="0D2B1A")
    sell_fill    = PatternFill("solid", fgColor="2B0D0D")
    border_side  = Side(style="thin", color="30363D")
    cell_border  = Border(bottom=Border(bottom=border_side).bottom)
    wrap         = Alignment(wrap_text=True, vertical="top")
    center       = Alignment(horizontal="center", vertical="top")

    # Write headers
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I1"

    # Collect and sort all trades
    all_trades = []
    for name, pf in portfolios.items():
        for t in pf.get("trades", []):
            all_trades.append({**t, "agent": name})
    all_trades.sort(key=lambda x: x["time"])

    for row_idx, t in enumerate(all_trades, 2):
        is_buy  = t["action"] == "BUY"
        row_fill = buy_fill if is_buy else sell_fill
        agent_color = AGENT_COLORS_HEX.get(t["agent"], "FFFFFF")

        vals = [
            t["time"][:16].replace("T", " "),
            t["agent"],
            t["action"],
            t["symbol"],
            int(t["shares"]),
            round(t["price"], 2),
            round(t["total"], 2),
            t.get("trade_pnl", ""),
            t.get("reason", ""),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.font = Font(color="E6EDF3", size=9)
            cell.alignment = wrap if col == 9 else center

        # Colour the agent name cell with the agent's own colour
        agent_cell = ws.cell(row=row_idx, column=2)
        agent_cell.font = Font(color=agent_color, bold=True, size=9)

        # Colour P&L cell green/red
        pnl_cell = ws.cell(row=row_idx, column=8)
        if isinstance(vals[7], (int, float)):
            pnl_cell.font = Font(
                color="3FB950" if vals[7] >= 0 else "F85149",
                bold=True, size=9
            )

        ws.row_dimensions[row_idx].height = 40

    # ── Sheet 2: Agent Summary ───────────────────────────────────────────
    ws2 = wb.create_sheet("Agent Summary")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 14
    ws2.column_dimensions["F"].width = 10

    sum_headers = ["Agent", "Portfolio Value", "Cash", "Started With", "P&L ($)", "P&L (%)"]
    for col, h in enumerate(sum_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 20
    ws2.freeze_panes = "A2"

    for row_idx, name in enumerate(AGENT_NAMES, 2):
        pf = portfolios.get(name, {})
        if not pf:
            ws2.cell(row=row_idx, column=1, value=name)
            ws2.cell(row=row_idx, column=2, value="No data yet")
            continue

        cash    = pf["cash"]
        start   = pf["starting_capital"]
        pos_val = sum(p["shares"] * p["avg_price"] for p in pf.get("positions", {}).values())
        total   = cash + pos_val
        pnl     = total - start
        pnl_pct = (pnl / start) * 100

        row_data = [name, round(total, 2), round(cash, 2), round(start, 2), round(pnl, 2), round(pnl_pct, 2)]
        agent_hex = AGENT_COLORS_HEX.get(name, "FFFFFF")

        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor="161B22")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == 1:
                cell.font = Font(color=agent_hex, bold=True, size=10)
            elif col in (5, 6):
                cell.font = Font(
                    color="3FB950" if val >= 0 else "F85149",
                    bold=True, size=10
                )
            else:
                cell.font = Font(color="E6EDF3", size=10)
        ws2.row_dimensions[row_idx].height = 22

    ws2.cell(row=len(AGENT_NAMES) + 3, column=1,
             value=f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}").font = Font(color="8B949E", size=9)

    wb.save(XLSX_PATH)
    print(f"Excel file rebuilt → {XLSX_PATH}")


# ─────────────────────────────────────────────────────────────────────────────
# HTML dashboard
# ─────────────────────────────────────────────────────────────────────────────

AGENT_DESCRIPTIONS = {
    "Momentum Agent":  "Ranks all 25 stocks by recent price gain and holds the top 5 with positive momentum. Rotates out when momentum fades.",
    "RSI Agent":       "Buys stocks that have been heavily sold off (RSI &lt; 35, oversold). Sells when the market gets too excited (RSI &gt; 65, overbought).",
    "MA Crossover":    "Buys when the short-term (10h) moving average crosses above the long-term (30h). Sells when it crosses back below.",
    "Dividend Agent":  "Focuses on high dividend-yield stocks (2%+) that are in a healthy uptrend — steady income while avoiding value traps.",
    "Sentiment Agent": "Reads recent Yahoo Finance news headlines and scores them. Buys positive coverage, sells negative news cycles.",
}


def compute_stats(portfolios: dict) -> dict:
    from collections import Counter
    all_trades = []
    for name, pf in portfolios.items():
        for t in pf.get("trades", []):
            all_trades.append({**t, "agent": name})

    sell_trades = [t for t in all_trades if t["action"] == "SELL" and "trade_pnl" in t]
    wins   = [t for t in sell_trades if t["trade_pnl"] > 0]
    win_rate = (len(wins) / len(sell_trades) * 100) if sell_trades else None

    best_trade  = max(sell_trades, key=lambda t: t["trade_pnl"])  if sell_trades else None
    worst_trade = min(sell_trades, key=lambda t: t["trade_pnl"])  if sell_trades else None
    most_traded = Counter(t["symbol"] for t in all_trades).most_common(1)[0] if all_trades else None

    total_pnl = 0
    total_value = 0
    best_agent = None
    best_pnl_pct = -9999

    agent_win_rates = {}
    for name, pf in portfolios.items():
        if not pf:
            agent_win_rates[name] = None
            continue
        start   = pf["starting_capital"]
        pos_val = sum(p["shares"] * p["avg_price"] for p in pf.get("positions", {}).values())
        total   = pf["cash"] + pos_val
        pnl_pct = (total - start) / start * 100
        total_pnl   += (total - start)
        total_value += total
        if pnl_pct > best_pnl_pct:
            best_pnl_pct = pnl_pct
            best_agent   = name

        agent_sells = [t for t in sell_trades if t["agent"] == name]
        agent_wins  = [t for t in agent_sells if t["trade_pnl"] > 0]
        agent_win_rates[name] = (len(agent_wins) / len(agent_sells) * 100) if agent_sells else None

    return dict(
        total_trades=len(all_trades),
        win_rate=win_rate,
        total_pnl=total_pnl,
        total_value=total_value,
        best_agent=best_agent,
        best_pnl_pct=best_pnl_pct,
        best_trade=best_trade,
        worst_trade=worst_trade,
        most_traded=most_traded,
        agent_win_rates=agent_win_rates,
    )


def build_overview(stats: dict, portfolios: dict) -> str:
    combined = f"${stats['total_value']:,.2f}" if stats['total_value'] else "—"
    pnl_sign = "+" if stats["total_pnl"] >= 0 else ""
    pnl_cls  = "pos" if stats["total_pnl"] >= 0 else "neg"
    combined_pnl = f'<span class="{pnl_cls}">{pnl_sign}${stats["total_pnl"]:,.2f}</span>'

    win_html = f"{stats['win_rate']:.1f}%" if stats["win_rate"] is not None else "No closed trades yet"
    best_agent_html = f"{stats['best_agent']} ({stats['best_pnl_pct']:+.2f}%)" if stats["best_agent"] else "—"

    bt = stats["best_trade"]
    wt = stats["worst_trade"]
    best_str  = f'<span class="pos">+${bt["trade_pnl"]:.2f}</span> on {bt["symbol"]}' if bt else "—"
    worst_str = f'<span class="neg">${wt["trade_pnl"]:.2f}</span> on {wt["symbol"]}' if wt else "—"
    mt = stats["most_traded"]
    most_str  = f"{mt[0]} ({mt[1]} trades)" if mt else "—"

    return f"""
<div class="overview-grid">
  <div class="ov-card">
    <div class="ov-label">Combined Portfolio Value</div>
    <div class="ov-value">{combined}</div>
    <div class="ov-sub">Total P&amp;L: {combined_pnl}</div>
  </div>
  <div class="ov-card">
    <div class="ov-label">Best Performing Agent</div>
    <div class="ov-value" style="font-size:1.1rem">{best_agent_html}</div>
  </div>
  <div class="ov-card">
    <div class="ov-label">Overall Win Rate</div>
    <div class="ov-value">{win_html}</div>
    <div class="ov-sub">{stats['total_trades']} total trades executed</div>
  </div>
  <div class="ov-card">
    <div class="ov-label">Best Trade</div>
    <div class="ov-value" style="font-size:1rem">{best_str}</div>
    <div class="ov-sub">Worst: {worst_str}</div>
  </div>
  <div class="ov-card">
    <div class="ov-label">Most Traded Stock</div>
    <div class="ov-value" style="font-size:1.1rem">{most_str}</div>
  </div>
</div>"""


def build_agent_card(name: str, pf: dict, win_rate) -> str:
    color = COLORS.get(name, "#888")
    desc  = AGENT_DESCRIPTIONS.get(name, "")

    if not pf:
        return f'''<div class="agent-card" style="border-top:3px solid {color}">
            <div class="agent-name">{name}</div>
            <div class="agent-desc">{desc}</div>
            <div class="muted" style="margin-top:12px">No data yet</div>
        </div>'''

    cash    = pf["cash"]
    start   = pf["starting_capital"]
    pos_val = sum(p["shares"] * p["avg_price"] for p in pf.get("positions", {}).values())
    total   = cash + pos_val
    pnl     = total - start
    pnl_pct = (pnl / start) * 100
    sign    = "+" if pnl >= 0 else ""
    pc      = "pos" if pnl >= 0 else "neg"

    wr_html = ""
    if win_rate is not None:
        wr_html = f'''<div class="win-rate-row">
            <span class="ov-label">Win rate</span>
            <span class="{'pos' if win_rate >= 50 else 'neg'}">{win_rate:.0f}%</span>
        </div>
        <div class="win-bar-bg"><div class="win-bar" style="width:{min(win_rate,100):.0f}%;background:{color}"></div></div>'''
    else:
        wr_html = '<div class="ov-label" style="margin-top:8px">Win rate: no closed trades yet</div>'

    positions_html = "".join(
        f'<span class="tag" style="border-color:{color}33">{sym} × {int(pos["shares"])}</span>'
        for sym, pos in pf.get("positions", {}).items()
    ) or '<span class="muted">No open positions</span>'

    return f'''<div class="agent-card" style="border-top:3px solid {color}">
        <div class="agent-name" style="color:{color}">{name}</div>
        <div class="agent-desc">{desc}</div>
        <div class="agent-value">${total:,.2f}</div>
        <div class="pnl {pc}">{sign}${pnl:,.2f} ({sign}{pnl_pct:.2f}%)</div>
        <div class="agent-meta">
            <span>Cash: ${cash:,.2f}</span>
            <span>Started: ${start:,.2f}</span>
            <span>{len(pf.get("trades",[]))} trades</span>
        </div>
        {wr_html}
        <div class="positions">{positions_html}</div>
    </div>'''


def build_trade_rows(portfolios: dict) -> str:
    all_trades = []
    for name, pf in portfolios.items():
        for t in pf.get("trades", []):
            all_trades.append({**t, "agent": name})
    all_trades.sort(key=lambda x: x["time"], reverse=True)

    rows = ""
    for t in all_trades[:100]:
        ts  = t["time"][5:16].replace("T", " ")
        ac  = "buy" if t["action"] == "BUY" else "sell"
        color = COLORS.get(t["agent"], "#fff")
        pnl_cell = ""
        if t["action"] == "SELL" and "trade_pnl" in t:
            p  = t["trade_pnl"]
            pc = "pos" if p >= 0 else "neg"
            pnl_cell = f'<span class="{pc} bold">{("+" if p >= 0 else "")}{p:.2f}</span>'
        reason = t.get("reason", "—")
        rows += f"""<tr>
            <td class="muted mono">{ts}</td>
            <td style="color:{color};font-weight:600;font-size:.8rem">{t['agent']}</td>
            <td><span class="badge {ac}">{t['action']}</span></td>
            <td class="bold">{t['symbol']}</td>
            <td class="muted">{int(t['shares'])}</td>
            <td>${t['price']:.2f}</td>
            <td>${t['total']:,.2f}</td>
            <td>{pnl_cell}</td>
            <td class="reason-cell">{reason}</td>
        </tr>"""

    if not rows:
        rows = '<tr><td colspan="9" class="muted" style="text-align:center;padding:28px">No trades yet — waiting for market open</td></tr>'
    return rows


def build_chart_data(snapshots: list) -> str:
    if not snapshots:
        return json.dumps({"labels": [], "datasets": []})
    labels = [s["time"][5:] for s in snapshots]
    datasets = []
    for name in AGENT_NAMES:
        datasets.append({
            "label": name,
            "data": [s.get(name) for s in snapshots],
            "borderColor": COLORS[name],
            "backgroundColor": COLORS[name] + "18",
            "fill": False,
            "tension": 0.4,
            "pointRadius": 0,
            "pointHoverRadius": 4,
            "borderWidth": 2,
        })
    return json.dumps({"labels": labels, "datasets": datasets})


def generate():
    portfolios  = {name: load_portfolio(name) for name in AGENT_NAMES}
    save_snapshot(portfolios)
    update_csv(portfolios)
    rebuild_xlsx(portfolios)
    snapshots   = load_snapshots()
    stats       = compute_stats(portfolios)

    overview_html = build_overview(stats, portfolios)
    cards_html    = "".join(build_agent_card(n, portfolios[n], stats["agent_win_rates"].get(n)) for n in AGENT_NAMES)
    trade_rows    = build_trade_rows(portfolios)
    chart_data    = build_chart_data(snapshots)
    updated       = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Paper Trading Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
:root {{
  --bg: #0d1117; --surface: #161b22; --border: #30363d;
  --text: #e6edf3; --muted: #8b949e; --accent: #58a6ff;
  --green: #3fb950; --red: #f85149; --surface2: #21262d;
}}
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ background: var(--bg); color: var(--text); font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; padding: 0; }}

/* ── Top bar ── */
.topbar {{ background: var(--surface); border-bottom: 1px solid var(--border); padding: 14px 28px; display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 10px; position: sticky; top: 0; z-index: 100; }}
.topbar-left {{ display: flex; align-items: center; gap: 14px; }}
.logo {{ font-size: 1.1rem; font-weight: 700; }}
.market-badge {{ padding: 3px 10px; border-radius: 20px; font-size: .75rem; font-weight: 600; }}
.market-open {{ background: #0d2b1a; color: var(--green); border: 1px solid var(--green); }}
.market-closed {{ background: #2b0d0d; color: var(--red); border: 1px solid var(--red); }}
.topbar-right {{ display: flex; align-items: center; gap: 20px; font-size: .82rem; color: var(--muted); }}
.countdown-box {{ text-align: center; }}
.countdown-label {{ font-size: .68rem; text-transform: uppercase; letter-spacing: .05em; }}
.countdown-val {{ font-size: 1.3rem; font-weight: 700; color: var(--text); font-variant-numeric: tabular-nums; }}
.et-time {{ font-size: .8rem; }}

/* ── Main content ── */
.content {{ max-width: 1500px; margin: 0 auto; padding: 24px 24px 40px; }}

/* ── Overview ── */
.overview-grid {{ display: flex; gap: 12px; margin-bottom: 28px; flex-wrap: wrap; }}
.ov-card {{ background: var(--surface); border: 1px solid var(--border); border-radius: 10px; padding: 16px 20px; flex: 1; min-width: 170px; }}
.ov-label {{ font-size: .72rem; text-transform: uppercase; letter-spacing: .06em; color: var(--muted); margin-bottom: 6px; }}
.ov-value {{ font-size: 1.4rem; font-weight: 700; margin-bottom: 4px; }}
.ov-sub {{ font-size: .78rem; color: var(--muted); }}

/* ── Agent cards ── */
.section-title {{ font-size: .75rem; text-transform: uppercase; letter-spacing: .08em; color: var(--muted); margin-bottom: 14px; }}
.agents-grid {{ display: flex; gap: 14px; margin-bottom: 28px; flex-wrap: wrap; }}
.agent-card {{ background: var(--surface); border: 1px solid var(--border); border-radius: 10px; padding: 18px; flex: 1; min-width: 220px; }}
.agent-name {{ font-size: .82rem; font-weight: 700; text-transform: uppercase; letter-spacing: .05em; margin-bottom: 6px; }}
.agent-desc {{ font-size: .76rem; color: var(--muted); line-height: 1.5; margin-bottom: 14px; min-height: 48px; }}
.agent-value {{ font-size: 1.7rem; font-weight: 700; margin-bottom: 4px; }}
.pnl {{ font-size: .88rem; font-weight: 600; margin-bottom: 10px; }}
.agent-meta {{ display: flex; gap: 10px; font-size: .74rem; color: var(--muted); margin-bottom: 10px; flex-wrap: wrap; }}
.win-rate-row {{ display: flex; justify-content: space-between; font-size: .76rem; margin-bottom: 4px; }}
.win-bar-bg {{ background: var(--surface2); border-radius: 4px; height: 5px; margin-bottom: 10px; }}
.win-bar {{ height: 5px; border-radius: 4px; transition: width .5s; }}
.positions {{ display: flex; flex-wrap: wrap; gap: 5px; margin-top: 6px; }}
.tag {{ background: var(--surface2); border: 1px solid var(--border); border-radius: 4px; padding: 2px 8px; font-size: .74rem; }}

/* ── Chart ── */
.chart-box {{ background: var(--surface); border: 1px solid var(--border); border-radius: 10px; padding: 20px; margin-bottom: 28px; }}

/* ── Trades table ── */
.table-header {{ display: flex; justify-content: space-between; align-items: baseline; margin-bottom: 12px; flex-wrap: wrap; gap: 8px; }}
.dl-link {{ font-size: .8rem; color: var(--accent); text-decoration: none; }}
.dl-link:hover {{ text-decoration: underline; }}
table {{ width: 100%; border-collapse: collapse; background: var(--surface); border: 1px solid var(--border); border-radius: 10px; overflow: hidden; font-size: .81rem; }}
thead {{ position: sticky; top: 57px; z-index: 10; }}
th {{ background: var(--surface2); padding: 10px 12px; text-align: left; font-size: .72rem; color: var(--muted); text-transform: uppercase; letter-spacing: .05em; white-space: nowrap; }}
td {{ padding: 9px 12px; border-top: 1px solid var(--surface2); vertical-align: top; }}
tr:hover td {{ background: #ffffff06; }}
.badge {{ padding: 2px 8px; border-radius: 4px; font-size: .74rem; font-weight: 700; }}
.badge.buy {{ background: #0d2b1a; color: var(--green); }}
.badge.sell {{ background: #2b0d0d; color: var(--red); }}
.reason-cell {{ color: var(--muted); font-size: .76rem; max-width: 380px; line-height: 1.5; }}
.pos {{ color: var(--green); }} .neg {{ color: var(--red); }}
.bold {{ font-weight: 600; }} .muted {{ color: var(--muted); }}
.mono {{ font-variant-numeric: tabular-nums; }}
</style>
</head>
<body>

<!-- ── Top bar ── -->
<div class="topbar">
  <div class="topbar-left">
    <div class="logo">📈 Paper Trading</div>
    <div id="market-badge" class="market-badge market-closed">Market Closed</div>
    <div id="et-time" class="et-time">—</div>
  </div>
  <div class="topbar-right">
    <div>
      <div class="ov-label">Data fetched</div>
      <div style="color:var(--text);font-size:.82rem">{updated}</div>
    </div>
    <div>
      <div class="ov-label">Data delay</div>
      <div style="color:var(--text);font-size:.82rem">~15 min</div>
    </div>
    <div class="countdown-box">
      <div class="countdown-label">Next cycle in</div>
      <div class="countdown-val" id="countdown">—</div>
    </div>
  </div>
</div>

<!-- ── Main ── -->
<div class="content">

  <!-- Overview -->
  <p class="section-title" style="margin-top:4px">Overview</p>
  {overview_html}

  <!-- Agent cards -->
  <p class="section-title">Trading Agents</p>
  <div class="agents-grid">{cards_html}</div>

  <!-- Chart -->
  <div class="chart-box">
    <p class="section-title">Portfolio Value Over Time</p>
    <canvas id="chart" height="55"></canvas>
  </div>

  <!-- Trades -->
  <div class="table-header">
    <p class="section-title" style="margin:0">Trade History &amp; Reasoning</p>
    <a class="dl-link" href="https://github.com/Gigantatoll/paper-trading/raw/main/data/trades.xlsx">⬇ Download Excel</a>
  </div>
  <table>
    <thead><tr>
      <th>Time (UTC)</th><th>Agent</th><th>Action</th><th>Stock</th>
      <th>Shares</th><th>Price</th><th>Total</th><th>Trade P&amp;L</th>
      <th>Why the agent did this</th>
    </tr></thead>
    <tbody>{trade_rows}</tbody>
  </table>

</div><!-- /content -->

<script>
// ── Chart ──────────────────────────────────────────────────────────────
const cd = {chart_data};
new Chart(document.getElementById('chart').getContext('2d'), {{
  type: 'line',
  data: {{ labels: cd.labels, datasets: cd.datasets }},
  options: {{
    responsive: true,
    interaction: {{ mode: 'index', intersect: false }},
    plugins: {{ legend: {{ labels: {{ color: '#e6edf3', boxWidth: 12, padding: 16 }} }} }},
    scales: {{
      x: {{ ticks: {{ color: '#8b949e', maxTicksLimit: 12 }}, grid: {{ color: '#21262d' }} }},
      y: {{ ticks: {{ color: '#8b949e', callback: v => '$' + v.toLocaleString() }}, grid: {{ color: '#21262d' }} }}
    }}
  }}
}});

// ── Countdown to next 5-min UTC cycle ──────────────────────────────────
function updateCountdown() {{
  const now = new Date();
  const secs = now.getUTCMinutes() % 5 * 60 + now.getUTCSeconds();
  const left = 300 - secs;
  const m = Math.floor(left / 60);
  const s = left % 60;
  document.getElementById('countdown').textContent = m + ':' + String(s).padStart(2,'0');
}}
setInterval(updateCountdown, 1000);
updateCountdown();

// ── Market status + ET clock ────────────────────────────────────────────
function updateMarket() {{
  const now  = new Date();
  const opts = {{ timeZone: 'America/New_York', hour: '2-digit', minute: '2-digit', second: '2-digit', hour12: false }};
  const etStr = now.toLocaleTimeString('en-US', opts);
  document.getElementById('et-time').textContent = 'ET ' + etStr;

  const et  = new Date(now.toLocaleString('en-US', {{ timeZone: 'America/New_York' }}));
  const day = et.getDay();
  const min = et.getHours() * 60 + et.getMinutes();
  const open = day >= 1 && day <= 5 && min >= 570 && min < 960;
  const badge = document.getElementById('market-badge');
  badge.textContent = open ? '● Market Open' : '● Market Closed';
  badge.className   = 'market-badge ' + (open ? 'market-open' : 'market-closed');
}}
setInterval(updateMarket, 1000);
updateMarket();
</script>
</body>
</html>"""

    os.makedirs(DOCS_DIR, exist_ok=True)
    with open(os.path.join(DOCS_DIR, "index.html"), "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Dashboard generated → {DOCS_DIR}/index.html")


if __name__ == "__main__":
    generate()
